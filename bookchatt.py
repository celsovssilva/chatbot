import pywhatkit
import pyautogui
import time
import os
import openpyxl
import threading
import tkinter as tk
from tkinter import messagebox
from datetime import datetime, timedelta

executando = False
PROGRESSO_ARQUIVO = "progresso.txt"
CONTATOS_RESPONDIDOS_ARQUIVO = "contatos_respondidos.txt"

def enviar_mensagem_whatsapp_auto(numero, nome, estado, municipio, verificar_resposta=False):
    """
    Envia uma mensagem automática pelo WhatsApp para um número especificado.

    Args:
        numero (str): O número de telefone do destinatário.
        nome (str): O nome do destinatário.
        estado (str): O estado do destinatário.
        municipio (str): O município do destinatário.
        verificar_resposta (bool): Indica se deve verificar se o contato respondeu.
    """
    try:
        numero_formatado = f"+55{numero.replace('-', '').replace(' ', '')}"
        mensagem = (
            f"Olá *{nome}*, meu nome é *Felipe*, sou Diretor Pedagógico do DEPARTAMENTO DE ASSISTÊNCIA A EDUCAÇÃO (SAEB) do Estado de *{estado}*, "
            f"estou entrando em contato devido ao *seu bom desempenho e comprometimento com a Educação, "
            f"através do Município de {municipio}.* Você ainda atua como educador?"
        )

        print(f"Enviando mensagem para {nome} ({numero_formatado})...")
        pywhatkit.sendwhatmsg_instantly(numero_formatado, mensagem, wait_time=15, tab_close=False) #wait time reduzido

        print("Aguardando para enviar...")
        time.sleep(5)
        pyautogui.press("enter")
        print("Mensagem enviada com sucesso!")
        time.sleep(5)

        if verificar_resposta:
            if verificar_se_contato_respondeu(numero_formatado):
                mensagem_resposta = "O motivo do meu contato é a respeito de uma seletiva feita pela Secretaria de Educação, na qual os profissionais foram indicados com base no resultado final do processo seletivo! Consegue me atender agora bem rapidinho para te explicar sobre o projeto?"
                print(f"Contato {nome} respondeu. Enviando segunda mensagem...")
                pywhatkit.sendwhatmsg_instantly(numero_formatado, mensagem_resposta, wait_time=10, tab_close=False) #wait time reduzido
                time.sleep(5)
                pyautogui.press("enter")
                print("Segunda mensagem enviada!")
                salvar_contato_respondido(numero_formatado)  # Salva o número que respondeu
                return True
            else:
                print(f"Contato {nome} não respondeu.")
                return False
        return False

    except Exception as e:
        print(f"[ERRO] Falha ao enviar mensagem para {nome} ({numero}): {e}")
        return False

def ler_progresso(planilha_atual):
    """
    Lê o progresso de envio de mensagens de um arquivo.

    Args:
        planilha_atual (str): O nome da planilha atual.

    Returns:
        int: A linha da última mensagem enviada ou 5 se não houver progresso.
    """
    if os.path.exists(PROGRESSO_ARQUIVO):
        with open(PROGRESSO_ARQUIVO, "r") as f:
            linhas = f.readlines()
            if len(linhas) >= 2:
                nome_salvo = linhas[0].strip()
                linha_salva = int(linhas[1].strip())
                if nome_salvo == planilha_atual:
                    return linha_salva
    return 5

def salvar_progresso(planilha_atual, linha):
    """
    Salva o progresso de envio de mensagens em um arquivo.

    Args:
        planilha_atual (str): O nome da planilha atual.
        linha (int): A linha da última mensagem enviada.
    """
    with open(PROGRESSO_ARQUIVO, "w") as f:
        f.write(f"{planilha_atual}\n{linha}")

def zerar_progresso(planilha_atual):
    """
    Zera o progresso de envio de mensagens para uma planilha.

    Args:
        planilha_atual (str): O nome da planilha atual.
    """
    with open(PROGRESSO_ARQUIVO, "w") as f:
        f.write(f"{planilha_atual}\n5")
    messagebox.showinfo("Reiniciado", "Progresso zerado com sucesso.")

def iniciar_envio():
    """
    Inicia o envio de mensagens em uma thread separada.
    """
    global executando
    executando = True
    thread = threading.Thread(target=main)
    thread.start()

def parar_envio():
    """
    Para o envio de mensagens e exibe uma mensagem de confirmação.
    """
    global executando
    executando = False
    messagebox.showinfo("Parar", "Envio de mensagens interrompido.")

def verificar_se_contato_respondeu(numero):
    """
    Verifica se um contato respondeu à mensagem nos últimos 2 minutos.
    Esta é uma SIMULAÇÃO e não verifica mensagens reais do WhatsApp.

    Args:
        numero (str): O número de telefone do contato.

    Returns:
        bool: True se o contato respondeu (simulado), False caso contrário.
    """
    # Simulação: Verifica se o número está na lista de respondidos.
    contatos_respondidos = ler_contatos_respondidos()
    if numero in contatos_respondidos:
        return True
    else:
        return False #simula que não respondeu

def salvar_contato_respondido(numero):
    """
    Salva o número de telefone do contato que respondeu em um arquivo.
    """
    with open(CONTATOS_RESPONDIDOS_ARQUIVO, "a") as f:
        f.write(f"{numero}\n")

def ler_contatos_respondidos():
    """
    Lê os números de telefone dos contatos que responderam do arquivo.
    """
    contatos = []
    if os.path.exists(CONTATOS_RESPONDIDOS_ARQUIVO):
        with open(CONTATOS_RESPONDIDOS_ARQUIVO, "r") as f:
            for linha in f:
                contatos.append(linha.strip())
    return contatos

def main():
    """
    Função principal que lê os dados da planilha Excel e envia as mensagens.
    """
    global executando
    # Obtém a lista de arquivos Excel e CSV no mesmo diretório do script
    arquivos_excel = [arquivo for arquivo in os.listdir() if arquivo.endswith(('.xlsx', '.xls', '.csv'))]

    if not arquivos_excel:
        messagebox.showerror("Erro", "Nenhum arquivo Excel ou CSV encontrado na pasta atual.")
        return

    arquivo_excel = arquivos_excel[0] # Pega o primeiro arquivo da lista

    # INSTRUÇÃO IMPORTANTE PARA O USUÁRIO
    messagebox.showinfo("Atenção", "Por favor, abra o WhatsApp Web no seu navegador e deixe-o carregado antes de clicar em 'Iniciar'.")

    try:
        # Imprime o diretório atual e o nome do arquivo para depuração
        print(f"Diretório atual: {os.getcwd()}")
        print(f"Arquivo a ser aberto: {arquivo_excel}")

        if not os.path.exists(arquivo_excel):
            messagebox.showerror("Erro", f"Arquivo não encontrado: {arquivo_excel}")
            return

        if arquivo_excel.endswith('.csv'): # Adicionado tratamento para CSV
            import csv
            with open(arquivo_excel, 'r', encoding='utf-8') as csvfile:
                reader = csv.reader(csvfile)
                header = next(reader, None)  # Lê o cabeçalho, se existir
                if header:
                    header_lower = [h.lower().strip() for h in header]
                    numero_coluna = next((i for i, h in enumerate(header_lower) if "numero" in h or "número" in h), None)
                    nome_coluna = next((i for i, h in enumerate(header_lower) if "nome" in h), None)
                    estado_coluna = next((i for i, h in enumerate(header_lower) if "estado" in h), None)
                    municipio_coluna = next((i for i, h in enumerate(header_lower) if "município" in h or "municipio" in h), None)

                if None in (numero_coluna, nome_coluna, estado_coluna, municipio_coluna):
                    messagebox.showerror("Erro", "As colunas 'Número', 'Nome', 'Estado' ou 'Município' não foram encontradas no arquivo CSV.")
                    return

                linha_inicio = ler_progresso(arquivo_excel)
                dados = list(reader) # Carrega todos os dados do CSV

                for linha_num, linha in enumerate(dados[linha_inicio-1:], start=linha_inicio):
                    if not executando:
                        print("Envio interrompido pelo usuário.")
                        break
                    try:
                        numero = str(linha[numero_coluna])
                        nome = linha[nome_coluna]
                        estado = linha[estado_coluna]
                        municipio = linha[municipio_coluna]

                        if None in (numero, nome, estado, municipio):
                            print(f"Linha {linha_num+1} ignorada: dados incompletos.")
                            continue

                        enviar_mensagem_whatsapp_auto(numero, nome, estado, municipio, verificar_resposta=True)
                        salvar_progresso(arquivo_excel, linha_num + 1)
                    
                    except Exception as inner_e:
                        print(f"[ERRO] Erro ao processar linha {linha_num+1}: {inner_e}")
                        continue

        else: # Se for um arquivo Excel
            workbook = openpyxl.load_workbook(arquivo_excel)
            sheet = workbook.active

            numero_coluna = nome_coluna = estado_coluna = municipio_coluna = None

            # Começar a leitura dos cabeçalhos a partir da linha 5 (índice 4)
            for i, coluna in enumerate(sheet[4]):
                if coluna.value is not None:
                    coluna_lower = str(coluna.value).lower().strip()
                    if "numero" in coluna_lower or "número" in coluna_lower:
                        numero_coluna = i
                    elif "nome" in coluna_lower:
                        nome_coluna = i
                    elif "estado" in coluna_lower:
                        estado_coluna = i
                    elif "município" in coluna_lower or "municipio" in coluna_lower:
                        municipio_coluna = i

            if None in (numero_coluna, nome_coluna, estado_coluna, municipio_coluna):
                messagebox.showerror("Erro", "As colunas 'Número', 'Nome', 'Estado' ou 'Município' não foram encontradas na linha 5.")
                return

            linha_inicio = ler_progresso(arquivo_excel)

            for linha_num in range(linha_inicio, sheet.max_row + 1):
                if not executando:
                    print("Envio interrompido pelo usuário.")
                    break

                try:
                    numero = str(sheet.cell(row=linha_num, column=numero_coluna + 1).value)
                    nome = sheet.cell(row=linha_num, column=nome_coluna + 1).value
                    estado = sheet.cell(row=linha_num, column=estado_coluna + 1).value
                    municipio = sheet.cell(row=linha_num, column=municipio_coluna + 1).value

                    if None in (numero, nome, estado, municipio):
                        print(f"Linha {linha_num} ignorada: dados incompletos.")
                        continue

                    enviar_mensagem_whatsapp_auto(numero, nome, estado, municipio, verificar_resposta=True)
                    salvar_progresso(arquivo_excel, linha_num + 1)
                 

                except Exception as inner_e:
                    print(f"[ERRO] Erro ao processar linha {linha_num}: {inner_e}")
                    continue

        if executando:
            messagebox.showinfo("Concluído", "Envio de mensagens finalizado.")
        executando = False

    except Exception as e:
        messagebox.showerror("Erro ao ler a planilha", str(e))


janela = tk.Tk()
janela.title("Envio de WhatsApp Automático")

frame_botoes = tk.Frame(janela, bg="#ADD8E6", bd=0) 
frame_botoes.pack(side=tk.BOTTOM, pady=20) 

botao_iniciar = tk.Button(frame_botoes, text="Iniciar", width=10, command=iniciar_envio, bg="green", fg="white")
botao_iniciar.pack(side=tk.LEFT, padx=5)

botao_parar = tk.Button(frame_botoes, text="Parar", width=10, command=parar_envio, bg="red", fg="white")
botao_parar.pack(side=tk.LEFT, padx=5)

def recomecar_zero():
    arquivos_excel = [arquivo for arquivo in os.listdir() if arquivo.endswith(('.xlsx', '.xls', '.csv'))] # Adicionado suporte a CSV
    if arquivos_excel:
        zerar_progresso(arquivos_excel[0])
    else:
        messagebox.showerror("Erro", "Nenhuma planilha encontrada para reiniciar.")

botao_recomecar = tk.Button(frame_botoes, text="Recomeçar", width=10, command=recomecar_zero, bg="blue", fg="white")
botao_recomecar.pack(side=tk.LEFT, padx=5)

janela.mainloop()
